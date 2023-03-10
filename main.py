import sys

import speech_recognition as sr
import texttospeech as ts
import chatbot
import time
from playaudio import play
from playaudio import louder_sound

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

WAKE_WORD = 'hello rami'
WAKE_WORD_VARIATIONS = [
    "hello ram",
    "hello mommy",
    "hello romy",
    "hello run",
    "hello robi",
    "hello ron",
    "hiram"
]

timestamp = time.strftime("%Y-%m-%d %H:%M:%S")


def handle_command(text):
    try:
        if text is not None:
            response = chatbot.request(text)
            if response is not None:
                return response
    except:
        print("unknown word")
        ts.speak("i currently dont know how to respond to that")
        pass


def get_wake_word():
    with sr.Microphone() as source:
        r = sr.Recognizer()
        r.pause_threshold = 0.8
        r.energy_threshold = 10000
        r.dynamic_energy_threshold = True
        audio = r.listen(source)
        text = r.recognize_google(audio)
        return text.lower()


def test_assistant():
    # create a new workbook to store data in an Excel sheet
    wb = Workbook()
    ws = wb.active

    # add headers to the Excel sheet
    ws['A1'] = 'Timestamp'
    ws['B1'] = 'Wake word received'
    ws['C1'] = 'Transcribed Input'
    ws['D1'] = 'Bot Response'
    ws['E1'] = 'Intents Tag Triggered'
    ws['F1'] = 'Elapsed time to get wake word'
    ws['G1'] = 'Elapsed Time Before Response'
    ws['H1'] = 'total elapsed time'

    # set font for headers
    bold_font = Font(bold=True)
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font

    row = 2  # start writing data from row 2

    while True:
        try:
            print('speak now')

            # record audio from microphone
            with sr.Microphone() as source:
                start_time = time.monotonic()
                r = sr.Recognizer()
                r.pause_threshold = 0.8
                r.energy_threshold = 10000
                r.dynamic_energy_threshold = True

                audio = r.listen(source)

                # transcribe audio input
                text = r.recognize_google(audio)
                text = text.lower()
                print("wakeword received text: " + text)  #

                # check wake word
                if any(variation in text for variation in WAKE_WORD_VARIATIONS):
                    detect_time = time.monotonic()
                    wake_word_time = detect_time - start_time
                    print(f"Wake word detection time: {wake_word_time:.4f} seconds")
                    ws.cell(row=row, column=6, value=wake_word_time)  # excel data wake word time
                    ws.cell(row=row, column=2, value=text)  # excel data wake word

                    print('now listening')
                    play(louder_sound)

                    # listen for the command after wake word is detected
                    audio = r.listen(source)
                    text = r.recognize_google(audio)
                    text = text.lower()
                    print("Recieved command: " + text)
                    ws.cell(row=row, column=3, value=text)  # excel data input text

                    # generate a response from the chatbot
                    response = handle_command(text)
                    if response:
                        ts.speak(response)
                    ws.cell(row=row, column=4, value=response)  # excel data output text

                    # measure elapsed time between wake word and chatbot response
                    end_time = time.monotonic()
                    ints = chatbot.get_tag(text)

                    try:
                        print("tag triggered: " + ints[0]['intent'])
                        ws.cell(row=row, column=5, value=ints[0]['intent'])
                    except:
                        pass

                    response_time = end_time - detect_time
                    ws.cell(row=row, column=7, value=response_time)

                    loop_time = end_time - start_time  # Calculate elapsed time
                    print(f"Loop time: {loop_time:.4f} seconds")  # Print elapsed time to 4 decimal places
                    ws.cell(row=row, column=8, value=loop_time)

                    # record the timestamp of the conversation
                    ws.cell(row=row, column=1, value=timestamp)

                    row += 1  # increment the row number for the next data entry
                    wb.save('chatbotdata.xlsx')



        except sr.RequestError:
            print("Could not request results from Google Speech Recognition service")
        except sr.UnknownValueError:
            print("Unable to recognize speech")
        except sr.WaitTimeoutError:
            print("Timeout error while waiting for speech input")
        except KeyboardInterrupt:
            ts.speak('bye')
            ts.engine.stop()
            sys.exit()


test_assistant()
