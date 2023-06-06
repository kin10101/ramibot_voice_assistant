import sys
import time

import speech_recognition as sr
from openpyxl import Workbook
from openpyxl.styles import Font

import chatbot
import texttospeech as ts
from playaudio import wakeSound, endSound
from playaudio import play

WAKE_WORD = 'hello rami'
WAKE_WORD_VARIATIONS = [
    "hello ram",
    "hello mommy",
    "hello romy",
    "hello run",
    "hello robi",
    "hello ron",
    "hiram",
    "hey rami",
    "rami",
    "hey ronnie",
    "jeremy",
    "hi rami",
    "hi ronnie"

]


def handle_command(text, context): # has bugs
    try:
        if text is not None:
            response = chatbot.handle_request(text, context)
            if response is not None:
                return response
    except:
        print("unknown word")
        # ts.speak("i currently don't know how to respond to that")
        pass


def get_wake_word():
    with sr.Microphone() as source:
        r = sr.Recognizer()
        r.pause_threshold = 0.8
        r.energy_threshold = 10000
        r.dynamic_energy_threshold = True
        #r.adjust_for_ambient_noise(source, duration=0.5)
        audio = r.listen(source)
        text = r.recognize_google(audio)
        return text.lower()




def test_assistant():
    # create a new workbook to store data in an Excel sheet
    wb = Workbook()
    ws = wb.active

    # add headers to the Excel sheet
    ws['A1'] = 'Timestamp'
    ws['B1'] = 'Wake word received'
    ws['C1'] = 'Transcribed Input'
    ws['D1'] = 'Bot Response'
    ws['E1'] = 'Intents Tag Triggered'
    ws['F1'] = 'Elapsed time to get wake word'
    ws['G1'] = 'Elapsed Time Before Response'
    ws['H1'] = 'total elapsed time'

    # set font for headers
    bold_font = Font(bold=True)
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font

    row = 2  # start writing data from row 2

    while True:
        wakeword_detected = False
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        try:
            print('speak now')

            # record audio from microphone
            with sr.Microphone() as source:
                print("1")
                start_time = time.monotonic()
                r = sr.Recognizer()
                r.pause_threshold = 0.8
                r.energy_threshold = 10000
                r.operation_timeout = 5000
                r.dynamic_energy_threshold = True
                print("set")

                audio = r.listen(source)
                print("listening now")

                # transcribe audio input
                text = r.recognize_google(audio)
                text = text.lower()
                print("wakeword received text: " + text)  #

                # check wake word
                if any(variation in text for variation in WAKE_WORD_VARIATIONS):
                    wakeword_detected = True
                    detect_time = time.monotonic()
                    wake_word_time = detect_time - start_time
                    print(f"Wake word detection time: {wake_word_time:.4f} seconds")
                    ws.cell(row=row, column=6, value=wake_word_time)  # excel data wake word time
                    ws.cell(row=row, column=2, value=text)  # excel data wake word

                    print('now listening')
                    play(wakeSound)

                    # listen for the command after wake word is detected
                    audio = r.listen(source, timeout=5)
                    text = r.recognize_google(audio, language='english')
                    text = text.lower()
                    print("Recieved command: " + text)
                    ws.cell(row=row, column=3, value=text)  # excel data input text

                    context = []
                    # generate a response from the chatbot
                    response = handle_command(text, context)
                    if response:
                        ts.speak(response)
                    ws.cell(row=row, column=4, value=response)  # excel data output text

                    # measure elapsed time between wake word and chatbot response
                    end_time = time.monotonic()
                    ints = chatbot.get_tag(text)

                    try:
                        print("tag triggered: " + ints[0]['intent'])
                        ws.cell(row=row, column=5, value=ints[0]['intent'])
                    except:

                        pass

                    response_time = end_time - detect_time
                    ws.cell(row=row, column=7, value=response_time)

                    loop_time = end_time - start_time  # Calculate elapsed time
                    print(f"Loop time: {loop_time:.4f} seconds")  # Print elapsed time to 4 decimal places
                    ws.cell(row=row, column=8, value=loop_time)

                    # record the timestamp of the conversation
                    ws.cell(row=row, column=1, value=timestamp)

                    row += 1  # increment the row number for the next data entry
                    wb.save('chatbotdata.xlsx')

                    play(endSound)  # sound to indicate that the conversation is over

                    if not chatbot.predict_class(text):  # if the input is not in the intents list
                        ts.speak("i currently don't know how to respond to that")

        except sr.RequestError:
            print("Could not request results from google Speech Recognition service")
        except sr.UnknownValueError:
            if wakeword_detected is True:
                play(endSound) #sound to indicate that the wake word was not detected

            print("Unable to recognize speech")
        except sr.WaitTimeoutError:
            print("Timeout error while waiting for speech input")
        except KeyboardInterrupt:
            ts.speak('bye')
            ts.engine.stop()
            sys.exit()


test_assistant()
